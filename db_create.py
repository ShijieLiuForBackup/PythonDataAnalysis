import openpyxl 
import sqlite3


# create the database if it does not exist
connection = sqlite3.connect('databaseA2.db')
cursor = connection.cursor()

# create inspections table
sql_command_inspect = """create table if not exists inspections (
id integer primary key, 
activity_date date, 
employee_id varchar(20), 
facility_address varchar(30), 
facility_city varchar(20),
facility_id varchar(20), 
facility_name varchar(20), 
facility_state varchar(10), 
facility_zip varchar(10), 
grade char(1), 
owner_id varchar(10), 
owner_name varchar(50), 
pe_description varchar(50),
program_element_pe varchar(5), 
program_name varchar(40), 
program_status varchar(10),
record_id varchar(20), 
score integer, 
serial_number varchar(20), 
serial_code integer, 
service_description varchar(30) );"""

# create violations table
sql_command_violate = """create table if not exists violations (
id integer primary key, 
points integer, 
serial_number varchar(20), 
violation_code varchar(10), 
violation_description varchar(100), 
violation_status varchar(50) );"""


# execute command
cursor.execute(sql_command_inspect)
cursor.execute(sql_command_violate)

# open violations sheet
wb_inspection = openpyxl.load_workbook('data/inspections.xlsx')
sheet = wb_inspection.get_sheet_by_name('inspections')
max_rows = sheet.max_row

# put all inspections' data to the inspections table
for i in range(2, max_rows+1):
    col1 = sheet.cell(row=i, column = 1).value
    col2 = sheet.cell(row=i, column = 2).value
    col3 = sheet.cell(row=i, column = 3).value
    col4 = sheet.cell(row=i, column = 4).value
    col5 = sheet.cell(row=i, column = 5).value
    col6 = sheet.cell(row=i, column = 6).value
    col7 = sheet.cell(row=i, column = 7).value
    col8 = sheet.cell(row=i, column = 8).value
    col9 = sheet.cell(row=i, column = 9).value
    col10 = sheet.cell(row=i, column = 10).value
    col11 = sheet.cell(row=i, column = 11).value
    col12 = sheet.cell(row=i, column = 12).value
    col13 = sheet.cell(row=i, column = 13).value
    col14 = sheet.cell(row=i, column = 14).value
    col15 = sheet.cell(row=i, column = 15).value
    col16 = sheet.cell(row=i, column = 16).value
    col17 = sheet.cell(row=i, column = 17).value
    col18 = sheet.cell(row=i, column = 18).value
    col19 = sheet.cell(row=i, column = 19).value
    col20 = sheet.cell(row=i, column = 20).value
    
    cursor.execute('INSERT INTO inspections VALUES (NULL, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
                    (col1, col2, col3, col4, col5, col6, col7, col8, col9, col10, col11, col12, col13, col14, col15,
                      col16, col17, col18, col19, col20))
connection.commit()

# open violations sheet
workbook_violation = openpyxl.load_workbook('data/violations.xlsx')
sheet = workbook_violation.get_sheet_by_name('violations')
max_Row = sheet.max_row 

# put all violations' data to the violstions table
for i in range(2,max_Row+1):
    col1 = sheet.cell(row=i, column = 1).value
    col2 = sheet.cell(row=i, column = 2).value
    col3 = sheet.cell(row=i, column = 3).value
    col4 = sheet.cell(row=i, column = 4).value
    col5 = sheet.cell(row=i, column = 5).value
    cursor.execute('INSERT INTO violations VALUES (NULL, ?, ? ,? ,?, ?)',
                   (col1, col2, col3, col4, col5) )

connection.commit()

print("done")
connection.close()

