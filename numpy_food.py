#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Sep 26 20:04:27 2018

@author: Shijie Liu
"""

import sqlite3
import matplotlib.pyplot as plt
import re
import openpyxl

# Question 1
#connect database
connection = sqlite3.connect('databaseA2.db')
cursor = connection.cursor()

months = ['01','02','03','04','05','06','07','08','09','10','11','12']


# question 1  select query
average_vio_query = """select inspections.facility_zip, printf("%.2f", count(violations.id)/30.0) as count from inspections\
                join violations on inspections.serial_number = violations.serial_number where inspections.activity_date between\
                "2015-07-01" and "2017-12-30" group by inspections.facility_zip"""

# execute query and get result
violations_avg = cursor.execute(average_vio_query).fetchall()
for item in violations_avg:
    print(item)

########## start to Question 2 ##########
def collecte_output(average_vio_query):
    # return data -> related to each month data
    each_month_vios = [] 
    # run the query (input query)
    execute_the_query = cursor.execute(average_vio_query).fetchall()
    # extract zip from the result
    get_zip = execute_the_query[0][0]
    # collect the zip data
    select_query = """ select count(violations.id) as count from inspections join violations on inspections.serial_number = violations.serial_number\
                where inspections.facility_zip = ? group by strftime('%m', inspections.activity_date)"""
    get_zip_data = cursor.execute(select_query, (get_zip,) ).fetchall()
    #convert collection to array and return
    for month_vio in get_zip_data:
        each_month_vios.append(month_vio[0])
    return each_month_vios


## Question 2.1 highest total violations 
#select query for the postcode has hihest total violations
highest_query = """select zip,  max(heighest) as heighest_vio from ( select inspections.facility_zip as zip,\
                 count(violations.id) as heighest from inspections join violations on inspections.serial_number = violations.serial_number\
                 where inspections.activity_date group by inspections.facility_zip)"""
# call the collect_result function
highest_violation = collecte_output(highest_query)
### display the MatPlot
plt.plot(months, highest_violation)
plt.show()
#
#
# Question 2.2 greatest variance (difference)

#select query for the zipcode has highest variance violations
highest_query_variance = """select facility_zip, max(variance) as max_variance from( select facility_zip,  (max(count) - min(count)) as variance\
                 from( select inspections.facility_zip,  count(violations.id) as count from inspections join\
                 violations on inspections.serial_number = violations.serial_number group by\
                 strftime('%m', inspections.activity_date), inspections.facility_zip) group by facility_zip)"""

each_month_variance = collecte_output(highest_query_variance)

## display the MatPlot
plt.plot(months, each_month_variance)
plt.show()


## Question 2.3
month_times = []
select_month_query = """select month, count(month) from (select strftime('%Y', inspections.activity_date) as year,\
                    strftime('%m', inspections.activity_date) as month,count(strftime('%m', inspections.activity_date)) as count\
                    from inspections group by strftime('%Y', inspections.activity_date), strftime('%m', inspections.activity_date)\
                    order by strftime('%Y', inspections.activity_date)) group by month"""
month_times_execute = cursor.execute(select_month_query).fetchall()
for month in month_times_execute:
    month_times.append(month[1])

select_vios_query = """select strftime('%m', inspections.activity_date), count(violations.id) from inspections join violations on\
                        inspections.serial_number = violations.serial_number group by strftime('%m', inspections.activity_date)"""
select_vios_result = cursor.execute(select_vios_query).fetchall()

month_avg_vios = []
for i in range(0, len(month_times)):
    month_avg_vios.append(float(format(select_vios_result[i][1] / month_times[i], '.2f')))

### display the MatPlot
plt.plot(months, month_avg_vios)
plt.show()
#### end 2.3 ####


# Question 2.4
# record number of each month violations
mc_bk_per_month_array = []

# select query for both Mc and Bk in each month
select_query = """select strftime('%m', inspections.activity_date),count(strftime('%m', inspections.activity_date))\
                 from inspections join violations on inspections.serial_number = violations.serial_number where\
                 inspections.program_name = 'MCDONALDS' or  inspections.program_name = 'BURGER KING' group by\
                 strftime('%m', inspections.activity_date) """
mc_bk_query = cursor.execute(select_query).fetchall()

# convert collection into array
for item in mc_bk_query:
    mc_bk_per_month_array.append(item[1])


#### display the MatPlot
plt.plot(months, mc_bk_per_month_array)
plt.title()
plt.show()

# Question 3.1

select_query = """select distinct violation_code, violation_description from violations order by violation_code """
distinct_result=cursor.execute(select_query).fetchall()
for item in distinct_result:
    print(item)


## Question 3.2

workb = openpyxl.load_workbook('data/violations.xlsx')
sheet = workb.get_sheet_by_name('violations')
max_row = sheet.max_row
result_dic = {}
result_list = []
for i in range(2,max_row + 1):
    violation_description = sheet.cell(row=i,column=4).value
    regex_search_result = re.search(r"[Ff]ood", violation_description)
    if regex_search_result != None:
        violation_code = sheet.cell(row=i,column=3).value
        result_dic[violation_code] = violation_description
        
for k,v in result_dic.items():
    each_result = (k,v)
    result_list.append(each_result)

print(result_list) 

    
    
    




