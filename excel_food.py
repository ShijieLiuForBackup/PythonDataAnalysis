#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Sep 26 15:00:24 2018

@author: Shijie Liu
"""

import openpyxl
import sqlite3
import os.path

if(os.path.exists('ViolationTypes.xlsx') == False):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Violations Types"
    wb.save('ViolationTypes.xlsx')
    print('create the file')
else:
    print('The file already exists and do not execut this py file more than 1 time, will result in data duplication')

# connect to databse and collect data
connect_databaseA2 = sqlite3.connect("databaseA2.db")
cursor = connect_databaseA2.cursor()
result = cursor.execute('select violation_code, violation_description, count(violation_code) as count\
                        from violations group by violation_code order by violation_code').fetchall()
# count row of data
num_types = len(result)

#set total number of violations
total_violations = 0

# Open and set column width
wb = openpyxl.load_workbook('ViolationTypes.xlsx')
ws = wb.worksheets[0]
ws['A2'] = "Code"
ws['B1'] = "Total Violations"
ws['B2'] = "Description"
ws['C2'] = "Count"
ws.column_dimensions['B'].width = 70

# insert data to excel sheet and count total violations
for i in range(0, num_types):
    ws.append(result[i])
    total_violations += result[i][2]

# inset number of total violations
ws['C1'] = total_violations
print(total_violations)
    
wb.save('ViolationTypes.xlsx')

 
 
 
 
 
 
 
 